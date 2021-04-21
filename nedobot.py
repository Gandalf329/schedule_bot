

import telegram
from telegram.ext import Updater
from telegram.ext import CommandHandler
from telegram.ext import MessageHandler, Filters
import logging
import time
import datetime
from openpyxl import load_workbook


token = '' #add token
bot = telegram.Bot(token)
print(bot.get_me())

updater = Updater(token, use_context=True) #Create Updater
dispatcher = updater.dispatcher #Create disaptcher for guick access

logging.basicConfig(format='%(asctime)s-%(name)s-%(levelname)s-%(message)s',level=logging.INFO) #set up logging module

def start(update,context: telegram.ext.CallbackContext):
    context.bot.send_message(chat_id=update.effective_chat.id, text = "Let's go me...") #create function start

start_handler = CommandHandler('start', start)
dispatcher.add_handler(start_handler) # /start is called and function "start" is started



def alarm(context: telegram.ext.CallbackContext): # Send message 
    context.bot.send_message(chat_id=context.job.context, text='BEEP')

def help(update,context: telegram.ext.CallbackContext):
     context.bot.send_message(chat_id=update.effective_chat.id, text = "Чтобы установить таймер введите /set[second]./rasp <день>\n /today <№ недели>")

help_handler = CommandHandler('help', help)
dispatcher.add_handler(help_handler)

def set_timer(update,context: telegram.ext.CallbackContext):
    chat_id = update.effective_chat.id
    try:
        due = int(context.args[0])
        print(due)
        if due < 0:
            update.message.reply_text("You are wrong.")
            return
        else:
            update.message.reply_text("Timer set.")
            time.sleep(due)
            update.message.reply_text("БДЗЫЫЫЫНЬ .")                   
    except (IndexError, ValueError):
        update.message.reply_text('Write in this format: /set <время>')

set_handler = CommandHandler('set', set_timer)
dispatcher.add_handler(set_handler)

def schedule(update,context: telegram.ext.CallbackContext):
    chat_id = update.effective_chat.id
    filename = 'rasp.xlsx'
    result = ''
    i = 1
    wb = load_workbook(filename)
    sheet = wb['w'] # w - is name sheet
    weeks = ['B','C','D','E','F','G','H']
    due = int(context.args[0])
    print(due)
    number = due-1
    word = weeks[number]

    column = sheet[word]

    while i < len(column):
        result += str(column[i].value) + ";\n"
        
        i += 1
    update.message.reply_text(result)

schedule_handler = CommandHandler('rasp', schedule)
dispatcher.add_handler(schedule_handler)

def schedule_today(update,context: telegram.ext.CallbackContext):
    chat_id = update.effective_chat.id
    due = int(context.args[0])
    print(due)
    if due == 1:
        filename = 'rasp 1.xlsx'
        result = ''
        i = 1
        wb = load_workbook(filename)
        sheet = wb['w'] # w - is name sheet
        weeks = ['B','C','D','E','F','G','H']
        today = datetime.datetime.today().isoweekday()
        number = today-1
        word = weeks[number]

        column = sheet[word]

        while i < len(column):
            result += str(column[i].value) + ";\n"
        
            i += 1
        update.message.reply_text(result)
    else:
        filename = 'rasp 2.xlsx'
        result = ''
        i = 1
        wb = load_workbook(filename)
        sheet = wb['w'] # w - is name sheet
        weeks = ['B','C','D','E','F','G','H']
        today = datetime.datetime.today().isoweekday()
        number = today-1
        word = weeks[number]

        column = sheet[word]

        while i < len(column):
            result += str(column[i].value) + ";\n"
        
            i += 1
        update.message.reply_text(result)

today_handler = CommandHandler('today', schedule_today)
dispatcher.add_handler(today_handler)

def unknown(update, context: telegram.ext.CallbackContext):
    context.bot.send_message(chat_id=update.effective_chat.id, text="Sorry, I didn't understand that command.")

unknown_handler = MessageHandler(Filters.command, unknown)
dispatcher.add_handler(unknown_handler)

updater.start_polling() # LAUNCH